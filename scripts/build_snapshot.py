#!/usr/bin/env python3
"""Build the dashboard RAW snapshot from Supabase source tables."""

import argparse
import json
import os
import re
import tempfile
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import requests
from openpyxl import Workbook
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

from revenue_rules import (
    canonical_daily,
    canonical_totals,
    canonical_unit_cost,
    canonicalize_sales_rows,
    is_cancelled,
)


ROOT = Path(__file__).resolve().parents[1]
HTML = ROOT / "index.html"
NUMERIC_COLUMNS = {"sku": {15, 16, 17, 28}, "tonkho": {5, 6}, "dim_class": {11, 12, 13}}


def extract_json_assignment(source: str, name: str):
    match = re.search(rf"{re.escape(name)}=(.+?);\n", source)
    if not match:
        raise RuntimeError(f"Không tìm thấy {name} trong index.html")
    return json.loads(match.group(1))


def extract_single_quoted(source: str, name: str) -> str:
    match = re.search(rf"var {re.escape(name)}='([^']+)'", source)
    if not match:
        raise RuntimeError(f"Không tìm thấy {name} trong index.html")
    return match.group(1)


def fetch_table(base_url: str, headers: dict, table: str) -> list[dict]:
    rows, offset, page_size = [], 0, 1000
    session = requests.Session()
    session.mount(
        "https://",
        HTTPAdapter(
            max_retries=Retry(
                total=5,
                backoff_factor=1,
                status_forcelist=(429, 500, 502, 503, 504),
                allowed_methods=("GET",),
            )
        ),
    )
    while True:
        response = session.get(
            f"{base_url}/rest/v1/{table}",
            headers=headers,
            params={"select": "*", "order": "_id.asc", "limit": page_size, "offset": offset},
            timeout=90,
        )
        response.raise_for_status()
        page = response.json()
        rows.extend(page)
        if len(page) < page_size:
            return rows
        offset += page_size


def as_number(value):
    if value in (None, ""):
        return None
    try:
        number = float(value)
        return int(number) if number.is_integer() else number
    except (TypeError, ValueError):
        return value


def build_workbook(table_map: dict, tables: dict, ads_only: bool) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for table, meta in table_map.items():
        if (table == "ads_fact") != ads_only:
            continue
        sheet = workbook.create_sheet(meta["sheet"])
        header_row = meta["hr"]
        numeric = NUMERIC_COLUMNS.get(table, set())
        for column in meta["cols"]:
            sheet.cell(header_row, column["c"], column["orig"])
        for row_index, row in enumerate(tables[table], start=header_row + 1):
            for column in meta["cols"]:
                value = row.get(column["col"])
                if column["c"] in numeric:
                    value = as_number(value)
                sheet.cell(row_index, column["c"], value)
    return workbook


def run_engine(source: str, table_map: dict, tables: dict) -> dict:
    engine = extract_json_assignment(source, "window._PY_ENGINE")
    real = extract_json_assignment(source, "window._PY_REAL")
    with tempfile.TemporaryDirectory(prefix="se-snapshot-") as tmp:
        tmp_path = Path(tmp)
        main_path = tmp_path / "main.xlsx"
        ads_path = tmp_path / "Sleep Expert — Ads Import.xlsx"
        engine_path = tmp_path / "build_engine.py"
        build_workbook(table_map, tables, False).save(main_path)
        build_workbook(table_map, tables, True).save(ads_path)
        engine_path.write_text(
            engine.replace("/sessions/sweet-pensive-hopper/mnt/outputs/", f"{tmp}/")
            .replace("/sessions/sweet-pensive-hopper/mnt/Data App/", f"{tmp}/")
            .replace("/sessions/sweet-pensive-hopper/mnt/uploads/", f"{tmp}/"),
            encoding="utf-8",
        )
        real = re.sub(r'XL\s*=\s*"[^"]*"', f'XL="{main_path}"', real)
        real = (
            real.replace("/sessions/sweet-pensive-hopper/mnt/outputs/build_engine.py", str(engine_path))
            .replace("/sessions/sweet-pensive-hopper/mnt/outputs/", f"{tmp}/")
            .replace("/sessions/sweet-pensive-hopper/mnt/Data App/", f"{tmp}/")
        )
        namespace = {}
        exec(compile(real, "<snapshot-builder>", "exec"), namespace)
        return namespace["RAW"]


def add_cancel_rates(raw: dict, sales_rows: list[dict]) -> None:
    """Attach monthly cancelled/gross GMV for dashboard period comparisons."""
    gmv_by_month: dict[str, list[float]] = defaultdict(lambda: [0.0, 0.0])
    for row in sales_rows:
        ordered_at = str(row.get("ngay_dat_hang") or "").strip()
        if len(ordered_at) < 7:
            continue
        month = ordered_at[:7].replace("-", ".")
        try:
            gmv = float(row.get("tong_gia_ban_san_pham") or 0)
        except (TypeError, ValueError):
            gmv = 0.0
        cancelled = is_cancelled(row.get("trang_thai_don_hang"))
        gmv_by_month[month][1] += gmv
        if cancelled:
            gmv_by_month[month][0] += gmv

    raw["cancelGmvByM"] = {
        month: [round(values[0], 2), round(values[1], 2)]
        for month, values in sorted(gmv_by_month.items())
    }


def add_ads_cancel_rates(raw: dict, canonical_rows: list[dict], sku_rows: list[dict]) -> None:
    """Replace legacy Ads cancellation rates with original Shopee GMV rates."""
    sku_to_product = {
        str(row.get("sku") or "").strip(): str(row.get("id_ma_san_pham_shopee") or "").strip()
        for row in sku_rows
        if str(row.get("sku") or "").strip()
    }
    by_product: dict[str, list[float]] = defaultdict(lambda: [0.0, 0.0])
    shop = [0.0, 0.0]
    for row in canonical_rows:
        gmv = float(row.get("g") or 0)
        cancelled = bool(row.get("x"))
        shop[1] += gmv
        if cancelled:
            shop[0] += gmv
        product_id = sku_to_product.get(str(row.get("s") or "").strip(), "")
        if product_id:
            by_product[product_id][1] += gmv
            if cancelled:
                by_product[product_id][0] += gmv

    ads = raw.get("ads")
    if not isinstance(ads, dict):
        return
    ads["cancelShop"] = round(shop[0] / shop[1], 6) if shop[1] else 0
    for row in ads.get("prod") or []:
        if isinstance(row, dict):
            product_id = str(row.get("shop") or "").strip()
        else:
            product_id = str(row[0] if isinstance(row, list) and row else "").strip()
        values = by_product.get(product_id)
        if not values:
            continue
        rate = round(values[0] / values[1], 6) if values[1] else 0
        if isinstance(row, dict):
            row["cancel"] = rate
        elif len(row) > 6:
            row[6] = rate


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="Tính snapshot nhưng không ghi Supabase")
    parser.add_argument("--output", type=Path, help="Ghi payload JSON để kiểm thử local")
    args = parser.parse_args()

    source = HTML.read_text(encoding="utf-8")
    table_map = extract_json_assignment(source, "window._SUPA_MAP")
    base_url = os.getenv("SUPABASE_URL") or extract_single_quoted(source, "SUPA_URL")
    api_key = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
    if not api_key:
        if not args.dry_run:
            raise RuntimeError("Thiếu secret SUPABASE_SERVICE_ROLE_KEY")
        api_key = extract_single_quoted(source, "SUPA_KEY")
    headers = {"apikey": api_key, "Authorization": f"Bearer {api_key}"}

    print(f"Đọc song song {len(table_map)} bảng nguồn…", flush=True)
    with ThreadPoolExecutor(max_workers=min(4, len(table_map))) as pool:
        futures = {
            table: pool.submit(fetch_table, base_url, headers, table) for table in table_map
        }
        tables = {table: future.result() for table, future in futures.items()}
    print("Số dòng:", {table: len(rows) for table, rows in tables.items()}, flush=True)

    engine_tables = dict(tables)
    # One COGS truth for the whole app: the legacy engine reads Unit Cost (VND),
    # therefore feed it the canonical after-tax cost before calculating every
    # Overview/chart/drill-down/inventory aggregate. P&L already applies the
    # same Giá vốn (+VAT) -> Unit Cost fallback in the browser.
    engine_tables["sku"] = [
        {**row, "unit_cost_vnd": canonical_unit_cost(row)}
        for row in tables.get("sku", [])
    ]
    engine_sales, canonical_sales = canonicalize_sales_rows(tables.get("sales_fact", []))
    engine_tables["sales_fact"] = engine_sales
    raw = run_engine(source, table_map, engine_tables)
    raw["salesCanonical"] = canonical_sales
    raw["salesD"] = canonical_daily(canonical_sales)
    raw["revenueRule"] = {
        "version": "shopee-net-after-tax-v1",
        "source": "Shopee API / sales_fact",
        "definition": "GMV đơn chưa hủy - giảm giá seller - hoàn trả thực tế",
        "totals": canonical_totals(canonical_sales),
    }
    add_cancel_rates(raw, tables.get("sales_fact", []))
    add_ads_cancel_rates(raw, canonical_sales, tables.get("sku", []))
    if tables.get("tonkho"):
        raw["stockDate"] = datetime.now(ZoneInfo("Asia/Ho_Chi_Minh")).strftime("%d/%m/%Y")
        raw["stockSource"] = "Shopee API"
    payload = json.dumps(raw, ensure_ascii=False, separators=(",", ":"))
    print(f"Snapshot hoàn tất: {len(payload):,} bytes", flush=True)
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(payload, encoding="utf-8")
        print(f"Đã ghi snapshot kiểm thử: {args.output}", flush=True)
    if args.dry_run:
        return

    response = requests.post(
        f"{base_url}/rest/v1/snapshot",
        headers={
            **headers,
            "Content-Type": "application/json",
            "Prefer": "resolution=merge-duplicates,return=minimal",
        },
        params={"on_conflict": "slot"},
        json={
            "slot": "current",
            "payload": raw,
            "bytes": len(payload.encode("utf-8")),
            "uploaded_by": "backend-snapshot",
        },
        timeout=120,
    )
    response.raise_for_status()
    print("Đã cập nhật snapshot/current trên Supabase.", flush=True)


if __name__ == "__main__":
    main()
