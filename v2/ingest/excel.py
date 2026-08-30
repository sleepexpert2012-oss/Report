"""Đọc file Excel ở mức thô nhất: trả về đúng những gì có trong ô.

Cố ý KHÔNG diễn giải gì ở đây — không đoán kiểu, không bỏ dòng, không sửa
giá trị. Mọi phán xét thuộc về lớp kiểm tra (validate.py). Tách như vậy để khi
số liệu sai còn biết chắc là sai ở khâu đọc hay khâu diễn giải.
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# Các giá trị tuy có ký tự nhưng ý nghĩa là "trống". Chuỗi "None" xuất hiện
# thật trong file hiện tại do một lần xuất dữ liệu trước đó ghi ra chữ None.
RONG = {"", "none", "null", "nan", "-", "—", "n/a", "na"}


def la_rong(value) -> bool:
    if value is None:
        return True
    return str(value).strip().lower() in RONG


def doc_sheet(path: Path, ten_sheet: str, dong_tieu_de: int) -> Tuple[Dict[str, str], List[Tuple[int, Dict[str, object]]]]:
    """Trả về (tiêu_đề_theo_cột, danh_sách_dòng).

    tiêu_đề_theo_cột : {"A": "Mã sản phẩm", "B": "SKU", ...}
    danh_sách_dòng   : [(số_dòng_excel, {"A": giá trị, "B": giá trị, ...}), ...]
                       chỉ gồm các dòng SAU dòng tiêu đề.
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    if ten_sheet not in wb.sheetnames:
        wb.close()
        raise KeyError(f"Không có sheet {ten_sheet!r}. Sheet đang có: {', '.join(wb.sheetnames)}")

    ws = wb[ten_sheet]
    tieu_de: Dict[str, str] = {}
    dong: List[Tuple[int, Dict[str, object]]] = []

    # Ở chế độ read_only, openpyxl có thể trả về EmptyCell không có .row/.column_letter,
    # nên lấy số dòng/chữ cái cột theo vị trí thay vì tin vào thuộc tính của ô.
    for chi_so, row in enumerate(ws.iter_rows(), start=1):
        so_dong = next((c.row for c in row if getattr(c, "row", None)), chi_so)
        if so_dong < dong_tieu_de:
            continue
        o = {
            get_column_letter(i): c.value
            for i, c in enumerate(row, start=1)
            if not la_rong(getattr(c, "value", None))
        }
        if so_dong == dong_tieu_de:
            tieu_de = {letter: str(v).strip() for letter, v in o.items()}
            continue
        if o:  # bỏ qua dòng trống hoàn toàn
            dong.append((so_dong, o))

    wb.close()
    return tieu_de, dong
